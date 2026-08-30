"""Reusable primitives for seeding a folder of Excel workbooks for local testing.

This module knows nothing about orders, customers, or any other business
entity — only how to write the four *shapes* of file
``extractors/files/local_excel.py`` knows how to tell apart (see that
module's docstring for the seed / snapshot / incremental distinction):

- :func:`write_static_workbook` — a small reference file, unchanged run to run.
- :func:`write_overwritten_snapshot` — one file replaced in full each run.
- :func:`write_appended_snapshot` — one file grown in place each run.
- :func:`write_incremental_partitions` — a new, separately named file per period.

A project cloned from this template keeps this file as-is and replaces
``scripts/demo_dataset.py`` with its own fake data — see that module's
docstring for the checklist. Nothing here is specific to the shipped demo.
"""

from collections.abc import Callable
from datetime import date
from io import BytesIO
from typing import Any

import structlog
from openpyxl import Workbook, load_workbook

from core.local_files import LocalFolderClient

log = structlog.get_logger()


def workbook_bytes(header: list[str], rows: list[list[Any]]) -> bytes:
    """Serialise a header and rows into ``.xlsx`` bytes."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def read_rows(content: bytes) -> list[list[Any]]:
    """Read a workbook's data rows (no header) as plain lists."""
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    return [list(row) for row in rows[1:] if any(cell is not None for cell in row)]


def find_item(items: list[dict[str, Any]], name: str) -> dict[str, Any] | None:
    """Find an item by name (case-insensitive) in a :meth:`~LocalFolderClient.list_children` result."""
    for item in items:
        if item.get("name", "").lower() == name.lower():
            return item
    return None


def write_static_workbook(
    client: LocalFolderClient,
    folder_path: str,
    file_name: str,
    header: list[str],
    rows: list[list[Any]],
) -> str:
    """Write a small reference workbook, unchanged run to run.

    Use this for a "seed" file: static-ish lookup data a real source would
    rarely update. Determinism is the caller's job — pass the same ``rows``
    every call (no random generation) so re-seeding doesn't invent new
    "reference" data.

    Args:
        client: Where to write the workbook.
        folder_path: Destination folder, relative to the store root.
        file_name: Name to give the workbook.
        header: Column names, in order.
        rows: The fixed rows to write.

    Returns:
        ``file_name``, for the caller to collect into a summary list.
    """
    client.upload_file(folder_path, file_name, workbook_bytes(header, rows))
    log.info("wrote_static_workbook", file_name=file_name, rows=len(rows))
    return file_name


def write_overwritten_snapshot(
    client: LocalFolderClient,
    folder_path: str,
    file_name: str,
    header: list[str],
    row_builder: Callable[[int], list[Any]],
    count: int,
) -> str:
    """Overwrite a workbook in full with ``count`` freshly generated rows.

    Use this for a "snapshot" file a user replaces wholesale each time it
    updates, e.g. an exported report re-saved over itself.

    Args:
        client: Where to write the workbook.
        folder_path: Destination folder, relative to the store root.
        file_name: Name to give the workbook. Any existing file of this name
            is fully replaced, not merged with.
        header: Column names, in order.
        row_builder: Builds row ``i`` of ``count`` (0-indexed). Close over
            your own RNG instance if you want fresh values per call — see
            ``scripts/demo_dataset.py`` for the pattern.
        count: How many rows to generate.

    Returns:
        ``file_name``, for the caller to collect into a summary list.
    """
    rows = [row_builder(i) for i in range(count)]
    client.upload_file(folder_path, file_name, workbook_bytes(header, rows))
    log.info("wrote_overwritten_snapshot", file_name=file_name, rows=count)
    return file_name


def write_appended_snapshot(
    client: LocalFolderClient,
    folder_path: str,
    file_name: str,
    header: list[str],
    next_id: Callable[[list[list[Any]]], int],
    row_builder: Callable[[int], list[Any]],
    new_count: int,
) -> str:
    """Read an existing workbook (if any), append new rows, and write it back whole.

    Use this for a "snapshot" file a user grows in place, e.g. a running log
    someone keeps adding rows to in the same workbook.

    Args:
        client: Where to read the existing workbook (if any) and write the
            updated one.
        folder_path: Destination folder, relative to the store root.
        file_name: Name of the workbook to read and rewrite.
        header: Column names, in order.
        next_id: Given the existing rows (``[]`` on a first run), returns the
            id new rows should continue from.
        row_builder: Builds one new row for a given id.
        new_count: How many new rows to append this run.

    Returns:
        ``file_name``, for the caller to collect into a summary list.
    """
    existing_item = find_item(client.list_children(folder_path), file_name)
    existing_rows: list[list[Any]] = []
    start_id = next_id([])
    if existing_item is not None:
        existing_rows = read_rows(client.download(existing_item["id"]))
        start_id = next_id(existing_rows)

    appended_rows = [row_builder(start_id + i) for i in range(new_count)]
    client.upload_file(
        folder_path, file_name, workbook_bytes(header, existing_rows + appended_rows)
    )
    log.info(
        "wrote_appended_snapshot",
        file_name=file_name,
        appended=new_count,
        total=len(existing_rows) + new_count,
    )
    return file_name


def month_start(months_ago: int) -> date:
    """Return the first day of the month ``months_ago`` months before today."""
    today = date.today()
    month_index = today.month - 1 - months_ago
    year = today.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def write_incremental_partitions(
    client: LocalFolderClient,
    folder_path: str,
    header: list[str],
    file_name_for: Callable[[date], str],
    row_builder_for: Callable[[date], Callable[[int], list[Any]]],
    rows_per_partition: int,
    backfill_months: int,
) -> list[str]:
    """Add any missing monthly partition files, oldest first; never touch existing ones.

    Use this for a series of separately-named files a real source drops on a
    schedule, e.g. one export per month. Existing partitions are left
    untouched — a partition already in the folder should never be
    regenerated, the way a real month that already landed wouldn't be.

    Args:
        client: Where to read the existing folder listing and write new files.
        folder_path: Destination folder, relative to the store root.
        header: Column names, in order.
        file_name_for: Names a period's file, given its month-start date.
        row_builder_for: Given a period's month-start date, returns a
            function building row ``i`` of that period's file (0-indexed).
            Called once per period, so it's the place to seed a
            per-period-deterministic RNG — see ``scripts/demo_dataset.py``.
        rows_per_partition: Rows per partition file.
        backfill_months: How many months of history to ensure exist, counting
            the current month (1 means "just this month").

    Returns:
        The newly created file names, oldest first.
    """
    existing_names = {
        item.get("name", "").lower() for item in client.list_children(folder_path)
    }

    created = []
    for months_ago in range(backfill_months - 1, -1, -1):
        period = month_start(months_ago)
        file_name = file_name_for(period)
        if file_name.lower() in existing_names:
            continue

        build_row = row_builder_for(period)
        rows = [build_row(i) for i in range(rows_per_partition)]
        client.upload_file(folder_path, file_name, workbook_bytes(header, rows))
        log.info(
            "wrote_incremental_partition",
            file_name=file_name,
            rows=rows_per_partition,
        )
        created.append(file_name)

    return created
