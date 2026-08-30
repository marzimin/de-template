"""Extractors for Excel workbooks stored in a local directory.

Four extractors, one per kind of file the folder holds, each needing a
different ``load_mode`` downstream:

``seed``
    Static reference data that barely changes, e.g. ``seed_categories.xlsx``.
    Filenames start with ``seed_``. Re-extracted in full every run —
    ``load_mode: replace`` in ``cfg/config.yaml`` is correct and cheap, since
    these workbooks are small.

``snapshot``
    One workbook that gets overwritten or appended to in place, e.g.
    ``latest_orders.xlsx`` or ``customers.xlsx`` — the filename never changes,
    but its content is the *full current state* each time it is read. Records
    carry ``_source_modified`` (the file's mtime), so a staging model — or a
    human — can tell whether a given run actually picked up new content or
    re-read the same snapshot. ``load_mode: replace`` is correct here too:
    there is nothing to append, the file already is the union of everything
    so far.

``incremental``
    A new, separately named file dropped on a schedule, e.g.
    ``monthly_sales_2026-08-01.xlsx``. Each file is a distinct partition with
    no overlap with the others, so ``load_mode: append`` is correct — re-running
    the pipeline after a new file lands adds only that file's rows, not the
    ones already loaded (deduplicate in staging if a run is retried, same as
    any other append source).

Each kind gets its own extractor class below, matched against the folder by a
filename pattern, so a bug or a schema change in one does not touch the
others. Register each under ``sources:`` in ``cfg/config.yaml`` with its own
``target_table`` and ``load_mode`` — see the ``local_*`` entries there.

This reads a plain directory rather than a live service on purpose: it lets
the whole pipeline (extract → load → dbt → export) be exercised with no
network call and no external account, for local development, tests, and CI.
It is a demo/testing data source, not a production one — swap in a real
extractor (see ``extractors/api/example_api.py`` for the pattern) when you
have a real upstream. See ``docs/pipelines.md#local-dummy-data`` for the full
picture. Connection details live in :mod:`core.local_files`, shared with the
generators in ``scripts/seed_toolkit.py``/``scripts/demo_dataset.py``, which
produce one example of each kind so the whole thing is runnable immediately
after cloning.

**The four concrete classes below are the shipped example, not the
mechanism.** ``LocalExcelExtractor`` (the base class) and ``_rows_from_workbook``
are the reusable parts — keep them. ``LocalSeedExtractor``,
``LocalLatestOrdersExtractor``, ``LocalCustomersExtractor``, and
``LocalMonthlySalesExtractor`` exist to match the demo data in
``scripts/demo_dataset.py``; when you replace that file with your own fake
(or real) data, copy whichever of the four classes matches each new file's
kind, give it a ``FILE_PATTERN`` for your filename, and register it under
``sources:`` in ``cfg/config.yaml``.
"""

import io
import re
from typing import Any

import structlog
from openpyxl import load_workbook

from core.local_files import LocalFolderClient, default_folder_path
from extractors.base import BaseExtractor

log = structlog.get_logger()


def _rows_from_workbook(content: bytes, item: dict[str, Any]) -> list[dict[str, Any]]:
    """Read the first sheet of a workbook into records.

    The first row is treated as the header. Rows are read with
    ``data_only=True`` so a formula cell yields its last-calculated value
    rather than the formula text.

    Args:
        content: Raw ``.xlsx`` bytes.
        item: The drive item the bytes came from, as returned by
            :meth:`core.local_files.LocalFolderClient.list_children`.

    Returns:
        One dict per data row, keyed by the header row's cell values, plus
        ``_source_file`` and ``_source_modified`` so a row can be traced back
        to the workbook — and the moment in time — it came from.
    """
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []

    source_file = item.get("name", "")
    source_modified = item.get("lastModifiedDateTime")
    records = [
        {
            **dict(zip(header, row, strict=False)),
            "_source_file": source_file,
            "_source_modified": source_modified,
        }
        for row in rows
        if any(cell is not None for cell in row)
    ]
    return records


class LocalExcelExtractor(BaseExtractor):
    """Base class: reads whichever workbooks in one local folder match ``FILE_PATTERN``.

    Not registered directly — subclass it and set :attr:`FILE_PATTERN` for
    each distinct kind of file the folder holds. See the module docstring for
    the three kinds shipped with this template.

    Reads ``local_dummy_data.destination`` and ``local_dummy_data.folder_path``
    from ``cfg/config.yaml`` — no secrets involved, so nothing needs to be set
    in ``.env``.
    """

    #: Matched against each item's file name (not its full path).
    FILE_PATTERN: re.Pattern[str]

    def __init__(self, client: LocalFolderClient | None = None) -> None:
        """Build the extractor.

        Args:
            client: Reuse an existing client. Built from ``cfg/config.yaml``
                via :meth:`LocalFolderClient.from_config` when omitted — pass
                one in tests to point at a temporary directory.
        """
        self.client = client or LocalFolderClient.from_config()
        self.folder_path = default_folder_path()

    def extract(self) -> list[dict[str, Any]]:
        """Read and parse every workbook in the folder matching :attr:`FILE_PATTERN`.

        Returns:
            One dict per data row across every matching workbook.
        """
        items = self.client.list_children(self.folder_path)
        matches = [
            item
            for item in items
            if "file" in item and self.FILE_PATTERN.match(item.get("name", ""))
        ]

        records: list[dict[str, Any]] = []
        for item in matches:
            content = self.client.download(item["id"])
            records.extend(_rows_from_workbook(content, item))

        log.info(
            "extracted",
            extractor=type(self).__name__,
            files=len(matches),
            records=len(records),
        )
        return records


class LocalSeedExtractor(LocalExcelExtractor):
    """Static reference workbooks, e.g. ``seed_categories.xlsx``. Load in full every run."""

    FILE_PATTERN = re.compile(r"^seed_.*\.xlsx$", re.IGNORECASE)


class LocalLatestOrdersExtractor(LocalExcelExtractor):
    """The single ``latest_orders.xlsx`` snapshot, overwritten in place as orders change."""

    FILE_PATTERN = re.compile(r"^latest_orders\.xlsx$", re.IGNORECASE)


class LocalCustomersExtractor(LocalExcelExtractor):
    """The single ``customers.xlsx`` snapshot, appended to in place as customers are added."""

    FILE_PATTERN = re.compile(r"^customers\.xlsx$", re.IGNORECASE)


class LocalMonthlySalesExtractor(LocalExcelExtractor):
    """Every ``monthly_sales_YYYY-MM-DD.xlsx`` partition dropped in the folder so far."""

    FILE_PATTERN = re.compile(r"^monthly_sales_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE)
