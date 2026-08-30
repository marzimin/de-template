from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.local_files import LocalFolderClient
from scripts.seed_toolkit import (
    month_start,
    write_appended_snapshot,
    write_incremental_partitions,
    write_overwritten_snapshot,
    write_static_workbook,
)

FOLDER = "Sample Data"
HEADER = ["id", "value"]


@pytest.fixture
def client(tmp_path):
    return LocalFolderClient(base_dir=tmp_path)


def _rows(client, file_name):
    item = next(i for i in client.list_children(FOLDER) if i["name"] == file_name)
    workbook = load_workbook(BytesIO(client.download(item["id"])), read_only=True)
    return list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))


def test_write_static_workbook_is_exactly_the_given_rows(client):
    write_static_workbook(client, FOLDER, "ref.xlsx", HEADER, [[1, "a"], [2, "b"]])

    rows = _rows(client, "ref.xlsx")

    assert rows == [("id", "value"), (1, "a"), (2, "b")]


def test_write_overwritten_snapshot_replaces_not_merges(client):
    write_overwritten_snapshot(
        client, FOLDER, "snap.xlsx", HEADER, row_builder=lambda i: [i, "old"], count=3
    )
    write_overwritten_snapshot(
        client, FOLDER, "snap.xlsx", HEADER, row_builder=lambda i: [i, "new"], count=2
    )

    rows = _rows(client, "snap.xlsx")

    assert rows == [("id", "value"), (0, "new"), (1, "new")]


def test_write_appended_snapshot_continues_the_id_sequence(client):
    write_appended_snapshot(
        client,
        FOLDER,
        "log.xlsx",
        HEADER,
        next_id=lambda existing: len(existing) + 1,
        row_builder=lambda i: [i, f"row-{i}"],
        new_count=2,
    )
    write_appended_snapshot(
        client,
        FOLDER,
        "log.xlsx",
        HEADER,
        next_id=lambda existing: len(existing) + 1,
        row_builder=lambda i: [i, f"row-{i}"],
        new_count=1,
    )

    rows = _rows(client, "log.xlsx")

    assert rows == [("id", "value"), (1, "row-1"), (2, "row-2"), (3, "row-3")]


def test_write_appended_snapshot_creates_file_on_first_run(client):
    write_appended_snapshot(
        client,
        FOLDER,
        "log.xlsx",
        HEADER,
        next_id=lambda existing: 1,
        row_builder=lambda i: [i, "first"],
        new_count=1,
    )

    rows = _rows(client, "log.xlsx")

    assert rows == [("id", "value"), (1, "first")]


def test_write_incremental_partitions_skips_existing_and_returns_new_only(client):
    first = write_incremental_partitions(
        client,
        FOLDER,
        HEADER,
        file_name_for=lambda period: f"part_{period.isoformat()}.xlsx",
        row_builder_for=lambda period: lambda i: [i, period.isoformat()],
        rows_per_partition=2,
        backfill_months=3,
    )
    assert len(first) == 3

    client.upload_file(FOLDER, first[0], b"hand-edited")

    second = write_incremental_partitions(
        client,
        FOLDER,
        HEADER,
        file_name_for=lambda period: f"part_{period.isoformat()}.xlsx",
        row_builder_for=lambda period: lambda i: [i, period.isoformat()],
        rows_per_partition=2,
        backfill_months=3,
    )

    assert second == []
    untouched = next(i for i in client.list_children(FOLDER) if i["name"] == first[0])
    assert client.download(untouched["id"]) == b"hand-edited"


def test_month_start_rolls_over_year_boundary():
    reference = date.today()
    months_to_january = reference.month - 1
    result = month_start(months_to_january)

    assert result.month == 1
    assert result.day == 1
