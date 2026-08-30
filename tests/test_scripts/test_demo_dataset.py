from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.local_files import LocalFolderClient
from scripts.demo_dataset import SEED_WORKBOOKS, seed_all

FOLDER = "Sample Data"


@pytest.fixture
def client(tmp_path):
    return LocalFolderClient(base_dir=tmp_path)


def _rows(client, file_name):
    item = next(i for i in client.list_children(FOLDER) if i["name"] == file_name)
    workbook = load_workbook(BytesIO(client.download(item["id"])), read_only=True)
    return list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))


def test_seed_all_writes_every_kind(client):
    uploaded = seed_all(client, FOLDER, backfill_months=2)

    names = {item["name"] for item in client.list_children(FOLDER)}
    assert "seed_categories.xlsx" in names
    assert "seed_regions.xlsx" in names
    assert "latest_orders.xlsx" in names
    assert "customers.xlsx" in names
    assert sum(name.startswith("monthly_sales_") for name in names) == 2
    assert set(uploaded) <= names


def test_seed_all_skip_static_omits_reference_files(client):
    seed_all(client, FOLDER, skip_static=True, backfill_months=1)

    names = {item["name"] for item in client.list_children(FOLDER)}
    assert "seed_categories.xlsx" not in names
    assert "latest_orders.xlsx" in names


def test_static_workbooks_are_deterministic_across_runs(client):
    seed_all(client, FOLDER, backfill_months=1)
    first = client.download(
        next(
            i
            for i in client.list_children(FOLDER)
            if i["name"] == "seed_categories.xlsx"
        )["id"]
    )

    seed_all(client, FOLDER, backfill_months=1)
    second = client.download(
        next(
            i
            for i in client.list_children(FOLDER)
            if i["name"] == "seed_categories.xlsx"
        )["id"]
    )

    assert first == second
    assert set(SEED_WORKBOOKS) == {"seed_categories.xlsx", "seed_regions.xlsx"}


def test_latest_orders_overwrites_in_place(client):
    seed_all(client, FOLDER, latest_orders_rows=5, backfill_months=1)
    first_rows = _rows(client, "latest_orders.xlsx")

    seed_all(client, FOLDER, latest_orders_rows=5, skip_static=True, backfill_months=1)
    second_rows = _rows(client, "latest_orders.xlsx")

    assert len(first_rows) == len(second_rows) == 6  # header + 5 rows
    items = [
        i for i in client.list_children(FOLDER) if i["name"] == "latest_orders.xlsx"
    ]
    assert len(items) == 1


def test_customers_appends_and_continues_the_id_sequence(client):
    seed_all(client, FOLDER, customers_new_rows=3, skip_static=True, backfill_months=1)
    first_rows = _rows(client, "customers.xlsx")
    assert len(first_rows) == 4  # header + 3

    seed_all(client, FOLDER, customers_new_rows=2, skip_static=True, backfill_months=1)
    second_rows = _rows(client, "customers.xlsx")
    assert len(second_rows) == 6  # header + 3 + 2, nothing dropped

    ids = [row[0] for row in second_rows[1:]]
    assert ids == [
        "CUST-00001",
        "CUST-00002",
        "CUST-00003",
        "CUST-00004",
        "CUST-00005",
    ]


def test_monthly_sales_does_not_regenerate_existing_months(client):
    seed_all(client, FOLDER, skip_static=True, backfill_months=3)
    monthly_files = sorted(
        item["name"]
        for item in client.list_children(FOLDER)
        if item["name"].startswith("monthly_sales_")
    )
    assert len(monthly_files) == 3

    client.upload_file(FOLDER, monthly_files[0], b"hand-edited")

    seed_all(client, FOLDER, skip_static=True, backfill_months=3)

    untouched = next(
        i for i in client.list_children(FOLDER) if i["name"] == monthly_files[0]
    )
    assert client.download(untouched["id"]) == b"hand-edited"
